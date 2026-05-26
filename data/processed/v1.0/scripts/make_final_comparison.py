from __future__ import annotations

import math
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from PIL import Image, ImageDraw, ImageFont


OUTPUT_FOLDER = "V1.0\u6570\u636e\u5904\u7406"
FINAL_FOLDER = "final_comparison"
CSV_ENCODING = "utf-8-sig"


def main() -> None:
    script_path = Path(__file__).resolve()
    output_root = script_path.parents[1]
    group_dir = output_root / "data" / "02_group_means"
    final_dir = output_root / FINAL_FOLDER
    final_dir.mkdir(parents=True, exist_ok=True)

    norm = pd.read_csv(group_dir / "all_group_means_smoothed_normalized_wide.csv")
    ref = pd.read_csv(group_dir / "all_group_means_reference_corrected_wide.csv")

    card_ids = sorted(
        [
            col[4:6]
            for col in norm.columns
            if col.startswith("card") and col.endswith("_MeanSmoothedNormalizedIntensity")
        ],
        key=int,
    )

    norm_final = pd.DataFrame({"Pixel": norm["Pixel"]})
    ref_final = pd.DataFrame({"Pixel": ref["Pixel"]})
    for card_id in card_ids:
        card_name = f"\u8272\u5361{int(card_id)}"
        norm_final[f"{card_name}_\u5e73\u6ed1\u5f52\u4e00\u5316\u5149\u5f3a"] = norm[
            f"card{card_id}_MeanSmoothedNormalizedIntensity"
        ]
        ref_final[f"{card_name}_\u53c2\u8003\u6821\u6b63\u76f8\u5bf9\u53cd\u5c04"] = ref[
            f"card{card_id}_MeanReferenceCorrectedRatio"
        ]

    features_df = make_feature_table(norm_final, ref_final, card_ids)
    rmse_df, maxdiff_df = make_pairwise_tables(norm_final, card_ids)

    norm_csv = final_dir / "V1.0_\u8272\u5361\u7ec4\u5747\u503c\u66f2\u7ebf_\u540c\u8868_\u5e73\u6ed1\u5f52\u4e00\u5316.csv"
    ref_csv = final_dir / "V1.0_\u8272\u5361\u7ec4\u5747\u503c\u66f2\u7ebf_\u540c\u8868_\u53c2\u8003\u6821\u6b63.csv"
    features_csv = final_dir / "V1.0_\u8272\u5361\u66f2\u7ebf\u5dee\u5f02\u6307\u6807.csv"
    norm_final.to_csv(norm_csv, index=False, encoding=CSV_ENCODING)
    ref_final.to_csv(ref_csv, index=False, encoding=CSV_ENCODING)
    features_df.to_csv(features_csv, index=False, encoding=CSV_ENCODING)

    norm_png = final_dir / "V1.0_\u8272\u5361\u6700\u7ec8\u5bf9\u6bd4\u56fe_\u5e73\u6ed1\u5f52\u4e00\u5316.png"
    ref_png = final_dir / "V1.0_\u8272\u5361\u6700\u7ec8\u5bf9\u6bd4\u56fe_\u53c2\u8003\u6821\u6b63\u76f8\u5bf9\u53cd\u5c04.png"
    draw_chart(
        norm_final,
        [col for col in norm_final.columns if col != "Pixel"],
        "V1.0 \u8272\u5361\u7ec4\u5747\u503c\u66f2\u7ebf\u5bf9\u6bd4 - \u5e73\u6ed1\u5f52\u4e00\u5316",
        "\u5e73\u6ed1\u5f52\u4e00\u5316\u5149\u5f3a",
        norm_png,
        y_min=0,
        y_max=1,
    )
    draw_chart(
        ref_final,
        [col for col in ref_final.columns if col != "Pixel"],
        "V1.0 \u8272\u5361\u7ec4\u5747\u503c\u66f2\u7ebf\u5bf9\u6bd4 - \u53c2\u8003\u6821\u6b63",
        "\u53c2\u8003\u6821\u6b63\u76f8\u5bf9\u53cd\u5c04",
        ref_png,
    )

    xlsx_path = final_dir / "V1.0_\u8272\u5361\u6700\u7ec8\u5bf9\u6bd4\u8868.xlsx"
    write_excel(xlsx_path, norm_final, ref_final, features_df, rmse_df, maxdiff_df)

    readme = f"""# V1.0 色卡最终对比结果

优先打开：

1. `{xlsx_path.name}`：所有色卡曲线在一个 Excel 文件里，包含曲线表、差异指标和图表。
2. `{norm_png.name}`：所有色卡平滑归一化曲线叠加图。
3. `{ref_png.name}`：所有色卡参考校正相对反射曲线叠加图。

CSV：

- `{norm_csv.name}`：Pixel + 每个色卡一列，适合看谱形差异。
- `{ref_csv.name}`：Pixel + 每个色卡一列，适合看参考校正后的相对反射差异。
- `{features_csv.name}`：峰值、面积、高低像素段比值等摘要指标。
"""
    (final_dir / "README.md").write_text(readme, encoding=CSV_ENCODING)

    print(f"Final comparison folder: {final_dir}")
    print(f"Excel: {xlsx_path}")
    print(f"Normalized PNG: {norm_png}")
    print(f"Reference-corrected PNG: {ref_png}")


def make_feature_table(norm_final: pd.DataFrame, ref_final: pd.DataFrame, card_ids: list[str]) -> pd.DataFrame:
    x = norm_final["Pixel"].to_numpy(float)
    low_region = (x >= 350) & (x <= 900)
    mid_region = (x >= 1200) & (x <= 2200)
    high_region = (x >= 2600) & (x <= 3500)

    rows = []
    for card_id in card_ids:
        card_name = f"\u8272\u5361{int(card_id)}"
        y_norm = norm_final[f"{card_name}_\u5e73\u6ed1\u5f52\u4e00\u5316\u5149\u5f3a"].to_numpy(float)
        y_ref = ref_final[f"{card_name}_\u53c2\u8003\u6821\u6b63\u76f8\u5bf9\u53cd\u5c04"].to_numpy(float)
        peak_idx = int(np.nanargmax(y_norm))
        ref_peak_idx = int(np.nanargmax(y_ref))
        low_mean = float(np.nanmean(y_ref[low_region]))
        high_mean = float(np.nanmean(y_ref[high_region]))
        rows.append(
            {
                "\u8272\u5361": int(card_id),
                "\u5f52\u4e00\u5316\u5cf0\u503c\u50cf\u7d20": int(x[peak_idx]),
                "\u5f52\u4e00\u5316\u66f2\u7ebf\u9762\u79ef": float(np.trapezoid(y_norm, x)),
                "\u53c2\u8003\u6821\u6b63\u5cf0\u503c\u50cf\u7d20": int(x[ref_peak_idx]),
                "\u53c2\u8003\u6821\u6b63\u6700\u5927\u503c": float(np.nanmax(y_ref)),
                "\u53c2\u8003\u6821\u6b63\u4e2d\u4f4d\u6570": float(np.nanmedian(y_ref)),
                "\u4f4e\u50cf\u7d20\u6bb5\u5747\u503c_350_900": low_mean,
                "\u4e2d\u50cf\u7d20\u6bb5\u5747\u503c_1200_2200": float(np.nanmean(y_ref[mid_region])),
                "\u9ad8\u50cf\u7d20\u6bb5\u5747\u503c_2600_3500": high_mean,
                "\u9ad8\u4f4e\u6bb5\u6bd4\u503c_2600_3500\u9664350_900": high_mean / low_mean,
            }
        )
    return pd.DataFrame(rows)


def make_pairwise_tables(norm_final: pd.DataFrame, card_ids: list[str]) -> tuple[pd.DataFrame, pd.DataFrame]:
    labels = [f"\u8272\u5361{int(card_id)}" for card_id in card_ids]
    rmse = pd.DataFrame(index=labels, columns=labels, dtype=float)
    maxdiff = pd.DataFrame(index=labels, columns=labels, dtype=float)
    for i, ci in enumerate(card_ids):
        yi = norm_final[f"\u8272\u5361{int(ci)}_\u5e73\u6ed1\u5f52\u4e00\u5316\u5149\u5f3a"].to_numpy(float)
        for j, cj in enumerate(card_ids):
            yj = norm_final[f"\u8272\u5361{int(cj)}_\u5e73\u6ed1\u5f52\u4e00\u5316\u5149\u5f3a"].to_numpy(float)
            diff = yi - yj
            rmse.iloc[i, j] = math.sqrt(float(np.nanmean(diff * diff)))
            maxdiff.iloc[i, j] = float(np.nanmax(np.abs(diff)))
    return rmse, maxdiff


def get_font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    candidates = [
        Path(r"C:\Windows\Fonts\msyhbd.ttc" if bold else r"C:\Windows\Fonts\msyh.ttc"),
        Path(r"C:\Windows\Fonts\simhei.ttf"),
        Path(r"C:\Windows\Fonts\arial.ttf"),
    ]
    for path in candidates:
        if path.exists():
            return ImageFont.truetype(str(path), size)
    return ImageFont.load_default()


def draw_chart(
    df: pd.DataFrame,
    columns: list[str],
    title: str,
    ylabel: str,
    path: Path,
    y_min: float | None = None,
    y_max: float | None = None,
) -> None:
    scale = 2
    width, height = 1800, 1040
    image = Image.new("RGB", (width * scale, height * scale), "white")
    draw = ImageDraw.Draw(image)

    title_font = get_font(34 * scale, True)
    label_font = get_font(22 * scale)
    tick_font = get_font(18 * scale)
    legend_font = get_font(20 * scale)
    palette = ["#1F77B4", "#D62728", "#2CA02C", "#9467BD", "#FF7F0E", "#17BECF", "#4D4D4D"]

    left, right, top, bottom = 130 * scale, 310 * scale, 95 * scale, 120 * scale
    plot_w = width * scale - left - right
    plot_h = height * scale - top - bottom
    x = df["Pixel"].to_numpy(float)
    ys = [df[col].to_numpy(float) for col in columns]
    x_min, x_max = float(np.nanmin(x)), float(np.nanmax(x))
    if y_min is None:
        y_min = float(np.nanmin(np.concatenate(ys)))
    if y_max is None:
        y_max = float(np.nanmax(np.concatenate(ys)))
    pad = (y_max - y_min) * 0.04 if y_max > y_min else 1
    y_min = max(0, y_min - pad)
    y_max = y_max + pad

    def sx(values: np.ndarray) -> np.ndarray:
        return left + (values - x_min) / (x_max - x_min) * plot_w

    def sy(values: np.ndarray) -> np.ndarray:
        return top + (y_max - values) / (y_max - y_min) * plot_h

    draw.text((width * scale / 2, 32 * scale), title, anchor="mm", fill="#111111", font=title_font)
    draw.rectangle([left, top, left + plot_w, top + plot_h], outline="#333333", width=2 * scale)

    for y_value in np.linspace(y_min, y_max, 7):
        y_pos = float(sy(np.array([y_value]))[0])
        draw.line([(left, y_pos), (left + plot_w, y_pos)], fill="#E4E4E4", width=1 * scale)
        draw.text((left - 14 * scale, y_pos), f"{y_value:.3f}", anchor="rm", fill="#444444", font=tick_font)

    for x_value in np.linspace(x_min, x_max, 7):
        x_pos = float(sx(np.array([x_value]))[0])
        draw.line([(x_pos, top), (x_pos, top + plot_h)], fill="#EFEFEF", width=1 * scale)
        draw.text((x_pos, top + plot_h + 28 * scale), f"{x_value:.0f}", anchor="mm", fill="#444444", font=tick_font)

    step = max(1, len(x) // 1400)
    for index, col in enumerate(columns):
        color = palette[index % len(palette)]
        y = df[col].to_numpy(float)
        finite = np.isfinite(y)
        points = list(zip(sx(x[finite][::step]).tolist(), sy(y[finite][::step]).tolist()))
        if len(points) > 1:
            draw.line(points, fill=color, width=4 * scale)

    draw.text((left + plot_w / 2, height * scale - 42 * scale), "Pixel", anchor="mm", fill="#111111", font=label_font)
    label_img = Image.new("RGBA", (460 * scale, 42 * scale), (255, 255, 255, 0))
    label_draw = ImageDraw.Draw(label_img)
    label_draw.text((230 * scale, 21 * scale), ylabel, anchor="mm", fill="#111111", font=label_font)
    label_img = label_img.rotate(90, expand=True)
    image.paste(label_img, (22 * scale, top + plot_h // 2 - label_img.height // 2), label_img)

    legend_x = left + plot_w + 42 * scale
    legend_y = top + 22 * scale
    for index, col in enumerate(columns):
        color = palette[index % len(palette)]
        label = col.split("_")[0]
        y_pos = legend_y + index * 42 * scale
        draw.line([(legend_x, y_pos), (legend_x + 44 * scale, y_pos)], fill=color, width=6 * scale)
        draw.text((legend_x + 58 * scale, y_pos), label, anchor="lm", fill="#111111", font=legend_font)

    image = image.resize((width, height), Image.Resampling.LANCZOS)
    image.save(path)


def write_excel(
    xlsx_path: Path,
    norm_final: pd.DataFrame,
    ref_final: pd.DataFrame,
    features_df: pd.DataFrame,
    rmse_df: pd.DataFrame,
    maxdiff_df: pd.DataFrame,
) -> None:
    wb = Workbook()
    ws_norm = wb.active
    ws_norm.title = "\u5e73\u6ed1\u5f52\u4e00\u5316\u66f2\u7ebf"
    append_dataframe(ws_norm, norm_final)

    ws_ref = wb.create_sheet("\u53c2\u8003\u6821\u6b63\u66f2\u7ebf")
    append_dataframe(ws_ref, ref_final)

    ws_features = wb.create_sheet("\u5dee\u5f02\u6307\u6807")
    append_dataframe(ws_features, features_df)

    ws_rmse = wb.create_sheet("\u4e24\u4e24RMSE")
    ws_rmse.append([""] + list(rmse_df.columns))
    for index, row in rmse_df.iterrows():
        ws_rmse.append([index] + [float(value) for value in row])

    ws_maxdiff = wb.create_sheet("\u4e24\u4e24\u6700\u5927\u5dee")
    ws_maxdiff.append([""] + list(maxdiff_df.columns))
    for index, row in maxdiff_df.iterrows():
        ws_maxdiff.append([index] + [float(value) for value in row])

    make_excel_chart(wb, ws_norm, "\u56fe_\u5e73\u6ed1\u5f52\u4e00\u5316", "V1.0 \u8272\u5361\u7ec4\u5747\u503c\u66f2\u7ebf\u5bf9\u6bd4 - \u5e73\u6ed1\u5f52\u4e00\u5316", "\u5e73\u6ed1\u5f52\u4e00\u5316\u5149\u5f3a")
    make_excel_chart(wb, ws_ref, "\u56fe_\u53c2\u8003\u6821\u6b63", "V1.0 \u8272\u5361\u7ec4\u5747\u503c\u66f2\u7ebf\u5bf9\u6bd4 - \u53c2\u8003\u6821\u6b63", "\u53c2\u8003\u6821\u6b63\u76f8\u5bf9\u53cd\u5c04")

    format_workbook(wb)
    wb.save(xlsx_path)


def append_dataframe(ws, df: pd.DataFrame) -> None:
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)


def make_excel_chart(wb: Workbook, source_ws, sheet_name: str, title: str, y_title: str) -> None:
    chart_ws = wb.create_sheet(sheet_name)
    chart_ws["A1"] = title
    chart_ws["A1"].font = Font(bold=True, size=14)

    chart = LineChart()
    chart.title = title
    chart.y_axis.title = y_title
    chart.x_axis.title = "Pixel"
    chart.height = 18
    chart.width = 32
    chart.legend.position = "r"
    chart.x_axis.majorGridlines = ChartLines()
    chart.y_axis.majorGridlines = ChartLines()

    data = Reference(source_ws, min_col=2, max_col=source_ws.max_column, min_row=1, max_row=source_ws.max_row)
    cats = Reference(source_ws, min_col=1, min_row=2, max_row=source_ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart_ws.add_chart(chart, "A3")


def format_workbook(wb: Workbook) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")

    for ws in wb.worksheets:
        if ws.max_row >= 1:
            ws.freeze_panes = "B2"
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
        for row in ws.iter_rows():
            for cell in row:
                cell.border = Border(bottom=thin)
        for col_idx in range(1, min(ws.max_column, 16) + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 14 if col_idx == 1 else 24


if __name__ == "__main__":
    main()
